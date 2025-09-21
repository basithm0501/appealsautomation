import csv
import sys
from copy import copy
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from openpyxl.utils import get_column_letter, column_index_from_string
import re

TEMPLATE = "AppealsTemplate.xlsx"

def create_appeals_workbook(csv_path, start_row, end_row=None):
    csv_path = convert_data_encoding(csv_path)
    wb = openpyxl.Workbook()
    print("✅ -- Workbook created. --")
    create_template_master_sheet(wb, "AppealsTemplate.xlsx")
    if end_row is None:
        end_row = start_row
    for row_num in range(start_row, end_row + 1):
        header, row = create_data_list(csv_path, row_num)
        create_appeal_sheet(wb, header, row)

    date_str = datetime.now().strftime("%Y-%m-%d")
    wb.save(f"CAP_Workbook_{date_str}.xlsx")


def convert_data_encoding(csv_path):
    with open(csv_path, encoding='latin1') as fin, open('AppealsData_utf8.csv', 'w', encoding='utf-8', newline='') as fout:
        for line in fin:
            fout.write(line)

    print(f"------------------------ NEW WORKBOOK ------------------------")
    print("✅ -- Appeals Data (ugly) converted to UTF-8 (pretty). --")
    return 'AppealsData_utf8.csv'


def create_template_master_sheet(wb, template_path):
    template_wb = openpyxl.load_workbook(template_path)
    template_ws = template_wb["master"]
    ws = wb.active
    ws.title = "master"

    # Copy all cells, including formatting, from template_ws to ws
    for row_idx, row in enumerate(template_ws.iter_rows(), 1):
        for col_idx, cell in enumerate(row, 1):
            new_cell = ws.cell(row=row_idx, column=col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    # Copy column widths
    for col_letter, col_dim in template_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = col_dim.width
    # Copy row heights
    for row_num, row_dim in template_ws.row_dimensions.items():
        ws.row_dimensions[row_num].height = row_dim.height

    print("✅ -- Template master sheet created. --")


def create_data_list(csv_path, row_num):
    with open(csv_path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        all_rows = list(reader)
    if row_num < 1 or row_num > len(all_rows):
        raise ValueError('Row number out of range')
    header_row = all_rows[2]  # Assuming the third row (index 2) is the header
    data_row = all_rows[row_num - 1]  # row_num is 1-based
    print(f"------------------------ NEW SHEET ------------------------")
    print(f"✅ -- Data processed for row: {row_num} --")
    return header_row, data_row


def create_appeal_sheet(wb, header, row):
    org_nickname = re.sub(r'[\\/*?:\[\]]', '_', row[11].strip().replace(" ", "_"))
    sheet_name = org_nickname[:31]
    ws = wb.create_sheet(title=sheet_name)
    copy_template_to_sheet(wb, ws, TEMPLATE)
    print(f"✅ -- Capsheet Created for: {org_nickname} --")

    # Get appeal1 and appeal2 by header index
    try:
        appeal1 = row[25]
        appeal2 = row[218]
    except ValueError:
        print("Header not found for appeal types.")
        appeal1 = appeal2 = None

    fill_header(ws, row)

    # Map funding type to fill function
    funding_type_to_func = {
        "Organizational Maintenance": fill_organizational_maintenance,
        "Stand Alone Program": fill_program1,
        "Series Program": fill_series_program,
        "Stand Alone Trip - Conference/Team Competition": fill_standalone_conference_team_competition,
        "Stand Alone Trip - Other": fill_standalone_other_trip,
        "Series Trip - Conference/Team Competition": fill_series_conference_team_competition, 
        "Series Trip - Other": fill_series_other_trip,
        "Magazine or Journal": fill_journal_magazine,
        "Newspaper": fill_newspaper,
    }

    # Helper to call the right function for each appeal
    def handle_appeal(appeal_type, appeal_num):
        func = funding_type_to_func.get(appeal_type)
        if func:
            func(ws, row, appeal_num)
        else:
            print(f"⚠️ No fill function mapped for: {appeal_type}")

    if appeal1 and appeal1 != "...":
        handle_appeal(appeal1, 1)
    if appeal2 and appeal2 not in ("...", "N/A", 'n/a', 'N/a', 'na', 'NA'):
        handle_appeal(appeal2, 2)

    fill_footer(ws, row)
    print(f"✅ -- Data inputted. --")


def copy_template_to_sheet(wb, ws, template_path):
    template_wb = openpyxl.load_workbook(template_path)
    # Use the "capsheet" sheet from the template workbook
    template_ws = template_wb["capsheet"]

    # Copy all cells, including formatting, from template_ws to ws
    for row_idx, row in enumerate(template_ws.iter_rows(), 1):
        for col_idx, cell in enumerate(row, 1):
            new_cell = ws.cell(row=row_idx, column=col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    # Copy column widths
    for col_letter, col_dim in template_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = col_dim.width
    # Copy row heights
    for row_num, row_dim in template_ws.row_dimensions.items():
        ws.row_dimensions[row_num].height = row_dim.height

# Done
def fill_header(ws, row):
    ws["A1"] = "Organization Name: " + (row[12])
    ws["A1"].font = Font(bold=True, size=16)
    ws["B1"] = "Submitted by: " + (row[16]) + " | Email: " + (row[17]) + " | Phone: " + str(row[18]) + " | Position: " + (row[19])
    ws["B1"].font = Font(bold=True)
    ws["K1"] = "SABO: " + str(row[14].strip().replace(" ", "").replace("-", "").replace("#", ""))
    ws["K1"].font = Font(bold=True, size=16)

# Done
def fill_program1(ws, data, appeal_num):
    if appeal_num == 1:
        print("----> First Appeal: Standalone Program")

        # Program Info Header
        ws["A2"] = f"Program 1 Name: {str(data[65]).strip()}"
        ws["A2"].font = Font(bold=True)
        ws["A3"] = f"Event Description: {str(data[66]).strip()}"
        ws["B3"] = f"Date: {str(data[67]).strip()}"
        ws["C3"] = f"Attendance: {str(data[69]).strip()}"
        ws["D3"] = f"Location: {str(data[70]).strip()}"
        ws["E3"] = f"Admission Fee: {str(data[71]).strip()}"

        # Room Rental and Equipment
        try:
            ws["B5"] = float(data[74].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B5"] = data[74]
        ws["C5"] = data[75]

        # Advertising
        try:
            ws["B6"] = float(data[76].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B6"] = data[76]
        ws["C6"] = data[77]

        # Food
        try:
            ws["B7"] = float(data[78].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B7"] = data[78]
        ws["C7"] = data[79]

        # Supplies + Duplications
        try:
            ws["B8"] = float(data[80].strip().replace("$", "").replace(",", "")) + float(data[82].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B8"] = f"{data[80]} + {data[82]}"
        ws["C8"] = data[81] + " | Duplications: " + data[83]

        # Contracts
        contract_indices = [84, 85, 86, 87, 88, 89, 90, 91]
        contracts = [data[i] for i in contract_indices]
        ws["C9"] = ", ".join(str(c) for c in contracts if str(c).strip())

        ws["B9"] = data[92]

        # Other
        try:
            ws["B17"] = float(data[93].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError): 
            ws["B17"] = data[93]
        ws["C17"] = data[94]

    else:
        fill_program2(ws, data, appeal_num)
    pass

# TODO
def fill_program2(ws, data, appeal_num):
    if appeal_num == 2:
        print("----> Second Appeal: Standalone Program")
    else:
        print("ERROR: fill_program2 called for appeal_num other than 2")
    pass

# Done
def fill_organizational_maintenance(ws, data, appeal_num):
    if appeal_num == 1:
        print("----> First Appeal: Organizational Maintenance")

        # Room Rental and Equipment + Storage
        try:
           ws["B24"] = float(data[28].strip().replace("$", "").replace(",", "")) + float(data[42].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
           if (str(data[28]).strip().replace("$", "").replace(",", "") not in ("", "0", "0.0", "N/A") or
               str(data[42]).strip().replace("$", "").replace(",", "") not in ("", "0", "0.0", "N/A")):
               ws["B24"] = str(data[28]) + " " + str(data[42])

        if str(data[42]).strip().replace("$", "").replace(",", "") not in ("", "0", "0.0", "N/A"):
            ws["C24"] = data[29] + " " + data[43]
        else:
            ws["C24"] = data[29]

        # Office Supplies
        try:
            ws["B25"] = float(data[30].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B25"] = data[30]
        ws["C25"] = data[31]

        # Advertising
        try:
            ws["B26"] = float(data[40].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B26"] = data[40]
        ws["C26"] = data[41]

        # Food for General Meetings
        try:
            ws["B27"] = float(data[44].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B27"] = data[44]
        ws["C27"] = data[45]

        # Promotional Giveaways
        try:
            ws["B28"] = float(data[32].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B28"] = data[32]
        ws["C28"] = data[33]

        # Software
        try:
            ws["B30"] = float(data[36].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B30"] = data[36]
        ws["C30"] = data[37]

        # Duplications
        try:
            ws["B31"] = float(data[38].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B31"] = data[38]
        ws["C31"] = data[39]

        # Phone Charges (Film Processing)
        try:
            ws["B32"] = float(data[34].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B32"] = data[34]
        ws["C32"] = data[35]

        # Uniforms/Costumes
        try:
            ws["B33"] = float(data[46].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B33"] = data[46]
        ws["C33"] = data[47]

        # Other
        try:
            ws["B35"] = float(data[48].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B35"] = data[48]
        ws["C35"] = data[49]

    else:
        print("----> Second Appeal: Organizational Maintenance")

        # Room Rental and Equipment + Storage (Room Rental Missing in Form)
        try:
            ws["B24"] = float(data[233].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B24"] = data[233]
        ws["C24"] = data[234]

        # Office Supplies
        try:
            ws["B25"] = float(data[221].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B25"] = data[221]
        ws["C25"] = data[222]

        # Advertising
        try:
            ws["B26"] = float(data[231].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B26"] = data[231]
        ws["C26"] = data[232]

        # Food for General Meetings
        try:
            ws["B27"] = float(data[235].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B27"] = data[235]
        ws["C27"] = data[236]

        # Promotional Giveaways
        try:
            ws["B28"] = float(data[223].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B28"] = data[223]
        ws["C28"] = data[224]

        # Software
        try:
            ws["B30"] = float(data[227].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B30"] = data[227]
        ws["C30"] = data[228]

        # Duplications
        try:
            ws["B31"] = float(data[229].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B31"] = data[229]
        ws["C31"] = data[230]

        # Phone Charges (Film Processing)
        try:
            ws["B32"] = float(data[225].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B32"] = data[225]
        ws["C32"] = data[226]

        # Uniforms/Costumes
        try:
            ws["B33"] = float(data[237].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B33"] = data[237]
        ws["C33"] = data[238]

        # Other
        try:
            ws["B35"] = float(data[239].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B35"] = data[239]
        ws["C35"] = data[240]

# TODO 2
def fill_series_program(ws, data, appeal_num):
    if appeal_num == 1:
        print("----> First Appeal: Series Program")

        # Program Info Header
        ws["G21"] = "Series Program Name: " + str(data[98]).strip()
        ws["J21"] = "Location: " + str(data[105]).strip()
        ws["K21"] = "Admissions Fee: " + str(data[106]).strip()
        ws["L21"] = "Installments: " + str(data[101]).strip()
        ws["G22"] = "Description: " + str(data[99]).strip()
        ws["K22"] = "Attendance: " + str(data[103]).strip()
        ws["L22"] = "Dates: " + str(data[102]).strip()

        # Room Rental and Equipment
        try:
            ws["H24"] = float(data[108].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H24"] = data[108]
        ws["I24"] = data[109]

        # Advertising
        try:
            ws["H25"] = float(data[110].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H25"] = data[110]
        ws["I25"] = data[111]

        # Food
        try:
            ws["H26"] = float(data[112].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H26"] = data[112]
        ws["I26"] = data[113]

        # Supplies + Duplications
        try:
            ws["H27"] = float(data[114].strip().replace("$", "").replace(",", "")) + float(data[116].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H28"] = f"{data[114]} + {data[116]}"
        ws["I27"] = "Supplies: " + data[115] + " | Duplications: " + data[117]

        # Contracts
        contract_indices = [118, 119, 120, 121, 122, 123, 124]
        contracts = [data[i] for i in contract_indices]
        ws["H29"] = ", ".join(str(c) for c in contracts if str(c).strip())
        ws["I28"] = data[125]

        ws["H28"] = data[126]

        # Other
        try:
            ws["H30"] = float(data[127].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H30"] = data[127]
        ws["I30"] = data[128]

    else:
        print("----> Second Appeal: Series Program")
    pass

# Done
def fill_journal_magazine(ws, data, appeal_num):
    if appeal_num == 1:
        print("----> First Appeal: Media Publication")

        # Publication Section Header
        ws["A38"] = "Media Publication (Journal/Magazine)"

        # Number of Issues
        try:
            ws["B42"] = int(data[52].strip().replace(",", ""))
        except (ValueError, TypeError):
            ws["B42"] = data[52]

        # Number of Pages per Issues
        try:
            ws["B44"] = int(data[53].strip().replace(",", ""))
        except (ValueError, TypeError):
            ws["B44"] = data[53]

        # Cost per Page
        ws["C43"] = f"Cost per page: {data[54]}"

        # Cost per Issue
        ws["C44"] = f"Cost per issue: {data[55]}"

        # Total Printing Costs
        if isinstance(ws["B42"].value, int) and isinstance(ws["B44"].value, int):
            try:
                ws["B45"] = ws["B42"].value * ws["B44"].value * float(data[54].strip().replace("$", "").replace(",", ""))
            except (ValueError, TypeError):
                ws["B45"] = f"{data[52]} x {data[53]} x {data[54]}"

        # Total Delivery Costs
        try:
            ws["B46"] = int(data[52].strip().replace(",", "")) * float(data[56].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B46"] = f"{data[52]} x {data[56]}"

    else:
        print("----> Second Appeal: Media Publication")

       # Publication Section Header
        ws["A38"] = "Media Publication (Journal/Magazine)"

        # Number of Issues
        try:
            ws["B42"] = int(data[242].strip().replace(",", ""))
        except (ValueError, TypeError):
            ws["B42"] = data[242]

        # Number of Pages per Issues
        try:
            ws["B44"] = int(data[243].strip().replace(",", ""))
        except (ValueError, TypeError):
            ws["B44"] = data[243]

        # Cost per Page
        ws["C43"] = f"Cost per page: {data[244]}"

        # Cost per Issue
        ws["C44"] = f"Cost per issue: {data[245]}"

        # Total Printing Costs
        if isinstance(ws["B42"].value, int) and isinstance(ws["B44"].value, int):
            try:
                ws["B45"] = ws["B42"].value * ws["B44"].value * float(data[244].strip().replace("$", "").replace(",", ""))
            except (ValueError, TypeError):
                ws["B45"] = f"{data[242]} x {data[243]} x {data[244]}"

        # Total Delivery Costs
        try:
            ws["B46"] = int(data[242].strip().replace(",", "")) * float(data[246].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B46"] = f"{data[242]} x {data[246]}"
            
    pass

# TODO 2
def fill_newspaper(ws, data, appeal_num):
    if appeal_num == 1:
        print("----> First Appeal: Media Publication")

        # Publication Section Header
        ws["A38"] = "Media Publication (Newspaper)"

        # Number of Issues
        try:
            ws["B42"] = int(data[59].strip().replace(",", ""))
        except (ValueError, TypeError):
            ws["B42"] = data[59]

        # Number of Pages per Issues
        try:
            ws["B44"] = int(data[60].strip().replace(",", ""))
        except (ValueError, TypeError):
            ws["B44"] = data[60]

        # Cost per Page
        ws["C43"] = f"Cost per page: {data[61]}"

        # Cost per Issue
        ws["C44"] = f"Cost per issue: {data[62]}"

        # Total Printing Costs
        if isinstance(ws["B42"].value, int) and isinstance(ws["B44"].value, int):
            try:
                ws["B45"] = ws["B42"].value * ws["B44"].value * float(data[61].strip().replace("$", "").replace(",", ""))
            except (ValueError, TypeError):
                ws["B45"] = f"{data[59]} x {data[60]} x {data[61]}"

        # Total Delivery Costs
        try:
            ws["B46"] = ws["B42"].value * float(data[63].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B46"] = f"{ws["B42"].value} x {data[63]}"

    else:
        print("----> Second Appeal: Media Publication")

        # Publication Section Header
        ws["A38"] = "Media Publication (Newspaper)"

        # Number of Issues
        try:
            ws["B42"] = int(data[250].strip().replace(",", ""))
        except (ValueError, TypeError):
            ws["B42"] = data[250]

        # Number of Pages per Issues
        try:
            ws["B44"] = int(data[251].strip().replace(",", ""))
        except (ValueError, TypeError):
            ws["B44"] = data[251]

        # Cost per Page
        ws["C43"] = f"Cost per page: {data[252]}"

        # Cost per Issue
        ws["C44"] = f"Cost per issue: {data[253]}"

        # Total Printing Costs
        if isinstance(ws["B42"].value, int) and isinstance(ws["B44"].value, int):
            try:
                ws["B45"] = ws["B42"].value * ws["B44"].value * float(data[252].strip().replace("$", "").replace(",", ""))
            except (ValueError, TypeError):
                ws["B45"] = f"{data[250]} x {data[251]} x {data[252]}"

        # Total Delivery Costs
        try:
            ws["B46"] = ws["B42"].value * float(data[254].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B46"] = f"{ws["B42"].value} x {data[254]}"

# TODO 2
def fill_standalone_other_trip(ws, data, appeal_num):
    if appeal_num == 1:
        print("----> First Appeal: Stand Alone Other Trip")

        # Header Information
        ws["G36"] = "Trip Name: " + str(data[153]).strip()
        ws["G36"].font = Font(bold=True)
        ws["H36"] = "Is this Series, If so Number of installments? NO"
        ws["H36"].font = Font(bold=True)
        ws["K36"] = "Location: " + str(data[159]).strip()
        ws["K36"].font = Font(bold=True)
        ws["H37"] = "Attendance: " + str(data[157]).strip()
        ws["H37"].font = Font(bold=True)
        ws["J37"] = f"Dates: {str(data[155]).strip()} - {str(data[156]).strip()}"
        ws["J37"].font = Font(bold=True)
        ws["G35"] = "Other Trip (Stand Alone | Series Trip) (Stand Alone)"
        ws["G35"].font = Font(bold=True)

        # Advertising
        try:
            ws["H39"] = float(data[161].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H39"] = data[161]
        ws["I39"] = data[162]

        # Transportation
        try:
            ws["H40"] = float(data[163].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H40"] = data[163]
        ws["I40"] = data[164]

        # Admission/Entry Fees
        try:
            ws["H41"] = float(data[165].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H41"] = data[165]
        ws["I41"] = data[166]

        # Food
        try:
            ws["H42"] = float(data[167].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H42"] = data[167]
        ws["I42"] = data[168]

        # Other
        try:
            ws["H44"] = float(data[169].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H44"] = data[169]
        ws["I44"] = data[170]

    else:
        print("----> Second Appeal: Stand Alone Other Trip")
    pass

# TODO 2
def fill_series_other_trip(ws, data, appeal_num):
    if appeal_num == 1:
        print("----> First Appeal: Series Other Trip")

        # Header Information
        ws["G36"] = "Trip Name: " + str(data[199]).strip()
        ws["G36"].font = Font(bold=True)
        ws["H36"] = "Is this Series, If so Number of installments? YES"
        ws["H36"].font = Font(bold=True)
        ws["K36"] = "Location: " + str(data[205]).strip()
        ws["K36"].font = Font(bold=True)
        ws["H37"] = "Attendance: " + str(data[203]).strip()
        ws["H37"].font = Font(bold=True)
        ws["J37"] = f"Dates: {str(data[201]).strip()} - {str(data[202]).strip()}"
        ws["J37"].font = Font(bold=True)
        ws["G35"] = "Other Trip (Stand Alone | Series Trip) (Series)"
        ws["G35"].font = Font(bold=True)

        # Advertising
        try:
            ws["H39"] = float(data[207].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H39"] = data[207]
        ws["I39"] = data[208]

        # Transportation
        try:
            ws["H40"] = float(data[209].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H40"] = data[209]
        ws["I40"] = data[210]

        # Admission/Entry Fees
        try:
            ws["H41"] = float(data[211].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H41"] = data[211]
        ws["I41"] = data[212]

        # Food
        try:
            ws["H42"] = float(data[213].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H42"] = data[213]
        ws["I42"] = data[214]

        # Other
        try:
            ws["H44"] = float(data[215].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["H44"] = data[215]
        ws["I44"] = data[216]

    else:
        print("----> Second Appeal: Series Other Trip")
    pass

# TODO 2
def fill_series_conference_team_competition(ws, data, appeal_num):
    if appeal_num == 1:
        print("----> First Appeal: Series Trip/Competition")

        # Header Information
        ws["A53"] = "Name of Competition/Conference: " + str(data[173]).strip()
        ws["A53"].font = Font(bold=True)
        ws["B53"] = "Is this Series, If so Number of installments? YES"
        ws["B53"].font = Font(bold=True)
        ws["B55"] = "Location: " + str(data[181]).strip()
        ws["B55"].font = Font(bold=True)
        ws["D55"] = "Attendance: " + str(data[177]).strip()
        ws["D55"].font = Font(bold=True)
        ws["F55"] = f"Dates: {str(data[175]).strip()} - {str(data[176]).strip()}"
        ws["F55"].font = Font(bold=True)
        ws["A52"] = f"Stand Alone Trip/Series - Competition/Conference  (max 6 trips in a series, max 1 series trip per semester) ({str(data[180]).strip()})"
        ws["A52"].font = Font(bold=True)

        # Transportation
        try:
            ws["B57"] = float(data[185].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B57"] = data[185]
        ws["C57"] = data[186]

        # Parking
        try:
            ws["B58"] = float(data[187].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B58"] = data[187]
        ws["C58"] = data[188]

        # Food
        try:
            ws["B59"] = float(data[189].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B59"] = data[189]
        ws["C59"] = data[190]

        # Lodging
        try:
            ws["B60"] = float(data[191].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B60"] = data[191]
        ws["C60"] = data[192]

        # Registration/Entry Fees
        try:
            ws["B61"] = float(data[193].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B61"] = data[193]
        ws["C61"] = data[194]

        # Other
        try:
            ws["B62"] = float(data[195].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B62"] = data[195]
        ws["C62"] = data[196]

    else:
        print("----> Second Appeal: Series Trip/Competition")
    pass

# TODO 2
def fill_standalone_conference_team_competition(ws, data, appeal_num):
    if appeal_num == 1:
        print("----> First Appeal: Standalone Conference/Team Competition")
        
        # Header Information
        ws["A53"] = "Name of Competition/Conference: " + str(data[130]).strip()
        ws["A53"].font = Font(bold=True)
        ws["B53"] = "Is this Series, If so Number of installments? NO"
        ws["B53"].font = Font(bold=True)
        ws["B55"] = "Location: " + str(data[138]).strip()
        ws["B55"].font = Font(bold=True)
        ws["D55"] = "Attendance: " + str(data[134]).strip()
        ws["D55"].font = Font(bold=True)
        ws["F55"] = f"Dates: {str(data[132]).strip()} - {str(data[133]).strip()}"
        ws["F55"].font = Font(bold=True)
        ws["A52"] = f"Stand Alone Trip/Series - Competition/Conference  (max 6 trips in a series, max 1 series trip per semester) ({str(data[137]).strip()})"
        ws["A52"].font = Font(bold=True)

        # Transportation
        try:
            ws["B57"] = float(data[140].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B57"] = data[140]
        ws["C57"] = data[141]

        # Parking
        try:
            ws["B58"] = float(data[142].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B58"] = data[142]
        ws["C58"] = data[143]

        # Food
        try:
            ws["B59"] = float(data[144].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B59"] = data[144]
        ws["C59"] = data[145]

        # Lodging
        try:
            ws["B60"] = float(data[146].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B60"] = data[146]
        ws["C60"] = data[147]

        # Registration/Entry Fees
        try:
            ws["B61"] = float(data[148].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B61"] = data[148]
        ws["C61"] = data[149]

        # Other
        try:
            ws["B62"] = float(data[150].strip().replace("$", "").replace(",", ""))
        except (ValueError, TypeError):
            ws["B62"] = data[150]
        ws["C62"] = data[151]

    else:
        print("----> Second Appeal: Standalone Conference/Team Competition")
    pass

# TODO
def fill_footer(ws, data):
    pass


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Usage: python AppealAutomation.py <AppealsData.csv> <start_row> <end_row>')
        sys.exit(1)
    csv_path = sys.argv[1]
    start_row = int(sys.argv[2])
    end_row = int(sys.argv[3]) if len(sys.argv) > 3 else None
    create_appeals_workbook(csv_path, start_row, end_row)
